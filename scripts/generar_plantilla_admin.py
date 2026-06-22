# -*- coding: utf-8 -*-
"""
Génère « PEEB - Plantilla datos Admin.xlsx » : un onglet par section/sous-section
de l'Admin, pré-rempli avec les données actuelles (lecture seule via l'API REST
Supabase en service_role). L'utilisateur complète / ajoute des lignes ; un import
ultérieur réinjecte dans la base.

Lance : python generar_plantilla_admin.py   (depuis la racine du repo)
"""
import os, json, ssl, urllib.request
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # racine du repo (depuis scripts/)
OUT = os.path.join(HERE, "PEEB - Plantilla datos Admin.xlsx")

# --- Lecture des identifiants depuis .env.local ----------------------------
def load_env(path):
    env = {}
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env

env = load_env(os.path.join(HERE, ".env.local"))
URL = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/")
KEY = env["SUPABASE_SERVICE_ROLE_KEY"]
CTX = ssl.create_default_context()

def fetch(table, params="select=*"):
    u = f"{URL}/rest/v1/{table}?{params}"
    req = urllib.request.Request(u, headers={
        "apikey": KEY, "Authorization": f"Bearer {KEY}", "Accept": "application/json",
    })
    with urllib.request.urlopen(req, context=CTX) as r:
        return json.loads(r.read().decode("utf-8"))

# --- Données ----------------------------------------------------------------
subs = fetch("peebcoolsf_subproyectos", "select=*&order=orden")
mets = fetch("peebcoolsf_metricas", "select=*")
meds = fetch("peebcoolsf_medidas", "select=*&order=subproyecto_uid,orden")
gest = fetch("peebcoolsf_gestion_lineas", "select=*&order=subproyecto_uid,orden")
equipo = fetch("peebcoolsf_equipo", "select=*&order=apellido,nombre")
ents = fetch("peebcoolsf_entidades", "select=*&order=nombre")
evts = fetch("peebcoolsf_eventos", "select=*&order=fecha")
gp = fetch("peebcoolsf_documentacion_gp", "select=*&order=orden")
cap = fetch("peebcoolsf_capacitaciones_documentos", "select=*&order=subseccion,orden")
fin = fetch("peebcoolsf_gestion_financiera", "select=*&order=uid")

sub_name = {s["uid"]: s["nombre"] for s in subs}
ent_name = {e["uid"]: e["nombre"] for e in ents}
def person_label(p):
    return (f'{(p.get("apellido") or "").strip()}, {(p.get("nombre") or "").strip()}').strip().strip(",").strip()
eq_name = {p["uid"]: person_label(p) for p in equipo}
part_label = {**eq_name, **ent_name}

MEDIDA_INFO = {
    "aislacion": ("Aislación", True), "carpinterias": ("Carpinterías", True),
    "hvac": ("HVAC", True), "luminarias": ("Luminarias", True),
    "fotovoltaicos": ("Fotovoltaicos", True), "solar_termica": ("Solar térmica", True),
    "genero": ("Género", False), "otras": ("Otras medidas", False),
    "ays": ("Ambiental y social", False),
}
FASE_NOMBRE = {
    "estudios_preliminares": "Estudios preliminares", "anteproyecto": "Anteproyecto",
    "proyecto_ejecutivo": "Proyecto ejecutivo", "redaccion_pliegos": "Redacción de pliegos",
    "no_objecion_afd": "No objeción AFD", "licitacion": "Licitación", "obra": "Obra",
    "general": "General",
}

def b(v):       # bool -> Sí/No
    return "Sí" if v else "No"
def hhmm(v):    # "HH:MM:SS" -> "HH:MM"
    return v[:5] if isinstance(v, str) and len(v) >= 5 else v

# --- Styles -----------------------------------------------------------------
HEAD_FILL = PatternFill("solid", fgColor="30323E")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
REF_FILL = PatternFill("solid", fgColor="EFEFEF")   # colonnes de référence (ne pas éditer)
NOTE_FONT = Font(italic=True, size=9, color="555555")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center")

ENUMS = {
    "comp": '"GP,EE,AyS,G"', "tip": '"A,H,E"',
    "secc": '"Aeropuertos,Hospitales,Escuelas"', "estado": '"en_proceso,terminado"',
    "modal": '"Presencial,Virtual"', "sexo": '"F,M,X"', "subsec": '"EE,AyS,G"',
    "bool": '"Sí,No"',
}

# col = (header, width, enum_key|None, ref?)
def add_sheet(wb, name, note, cols, rows, extra=0, gray_kwh_for=None):
    ws = wb.create_sheet(name[:31])
    n = len(cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(1, 1, note); c.font = NOTE_FONT; c.alignment = WRAP_TOP
    ws.row_dimensions[1].height = 46
    for j, col in enumerate(cols, 1):
        h = ws.cell(2, j, col[0]); h.font = HEAD_FONT; h.fill = HEAD_FILL; h.alignment = CENTER
        ws.column_dimensions[get_column_letter(j)].width = col[1]
    ws.row_dimensions[2].height = 30
    for i, row in enumerate(rows, 3):
        for j, col in enumerate(cols, 1):
            cell = ws.cell(i, j, row[j - 1])
            if col[3]:
                cell.fill = REF_FILL
            if gray_kwh_for and col[0].startswith("kWh") and not gray_kwh_for(row):
                cell.fill = REF_FILL
    # validations (sur les lignes de données + `extra` lignes vides pour les ajouts)
    last = 2 + len(rows) + extra
    for j, col in enumerate(cols, 1):
        if col[2]:
            dv = DataValidation(type="list", formula1=ENUMS[col[2]], allow_blank=True)
            ws.add_data_validation(dv)
            L = get_column_letter(j)
            dv.add(f"{L}3:{L}{max(last,3)}")
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = True
    return ws

wb = Workbook()
wb.remove(wb.active)

# === Instrucciones ==========================================================
ins = wb.create_sheet("Instrucciones")
ins.column_dimensions["A"].width = 110
lines = [
    ("PEEB Cool — Santa Fe · Plantilla de datos del Admin", 14, True),
    ("", 10, False),
    ("Una pestaña por sección / subsección del Admin. Completá los datos y devolveme el archivo; yo los importo a la app.", 10, False),
    ("", 6, False),
    ("Convenciones:", 11, True),
    ("• UID: las filas existentes ya traen su UID (columna gris). NO lo cambies ni borres esas filas.", 10, False),
    ("• Para AGREGAR (eventos, personas, entidades, documentos…): nueva fila al final, dejá UID vacío (yo lo genero).", 10, False),
    ("• Vacío = dato faltante. Nunca pongas 0 si no es un 0 real (se muestra « — » en la app).", 10, False),
    ("• Fechas: AAAA-MM-DD (ej. 2026-06-21). Horas: HH:MM (ej. 14:30).", 10, False),
    ("• Sí / No: usá el desplegable (confidencial, publicar, formación, activa…).", 10, False),
    ("• Listas (componente, tipología, estado, modalidad, sexo, subsección): usá el desplegable de la celda.", 10, False),
    ("• Componente: GP=Gestión de proyecto, EE=Eficiencia energética, AyS=Ambiental y social, G=Género.", 10, False),
    ("• Tipología: A=Aeropuertos, H=Hospitales, E=Escuelas.", 10, False),
    ("• Participantes (Eventos) y Entidad (Personas): escribí el NOMBRE tal cual figura en su pestaña; varios separados por « ; ».", 10, False),
    ("• Columnas grises = referencia (no editar): sirven para que sepas a qué fila corresponde.", 10, False),
    ("", 6, False),
    ("Pestañas FIJAS (no agregar ni borrar filas; solo completar valores):", 11, True),
    ("  Subproyectos · Factibilidad, Subproyectos · Proyecto, Subproyectos · Medidas, Subproyectos · Fases.", 10, False),
    ("  (En « Edificio » sí podés agregar una ESCUELA nueva: fila al final, UID vacío, Sección = Escuelas.)", 10, False),
    ("", 6, False),
    ("kWh/año (Medidas): solo aplica a las 6 medidas EE. Género, Otras y AyS no llevan kWh (celda gris).", 10, False),
]
for i, (txt, sz, bold) in enumerate(lines, 1):
    cc = ins.cell(i, 1, txt); cc.font = Font(size=sz, bold=bold); cc.alignment = WRAP_TOP

# === 1. GP — Documentación ==================================================
rows = [[g["uid"], g.get("nombre_documento"), g.get("componente"), g.get("url"),
         b(g.get("confidencial")), b(g.get("publicar"))] for g in gp]
add_sheet(wb, "GP - Documentación",
          "Documentación de proyecto. Podés agregar filas (UID vacío). Confidencial = oculto para Consultor; Publicar = visible en páginas públicas.",
          [("UID", 16, None, True), ("Documento", 40, None, False), ("Componente", 12, "comp", False),
           ("Enlace (URL)", 34, None, False), ("Confidencial", 12, "bool", False), ("Publicar", 12, "bool", False)],
          rows, extra=120)

# === 2. GP — Gestión financiera =============================================
rows = [[f["uid"], f.get("titulo"), f.get("url"), b(f.get("confidencial")), b(f.get("publicar"))] for f in fin]
add_sheet(wb, "GP - Gestión financiera",
          "Gestión financiera (estructura mínima por ahora: título + enlace). Podés agregar filas (UID vacío).",
          [("UID", 16, None, True), ("Título", 40, None, False), ("Enlace (URL)", 34, None, False),
           ("Confidencial", 12, "bool", False), ("Publicar", 12, "bool", False)],
          rows, extra=120)

# === 3. Calendario — Eventos ================================================
def parts(e):
    return "; ".join(part_label.get(u, u) for u in (e.get("participantes") or []))
rows = [[e["uid"], e.get("nombre"), e.get("fecha"), hhmm(e.get("hora_inicio")), hhmm(e.get("hora_fin")),
         e.get("componente"), e.get("modalidad"), e.get("lugar"), e.get("url_conexion"),
         parts(e), b(e.get("formacion")), e.get("url_documento")] for e in evts]
add_sheet(wb, "Calendario - Eventos",
          "Eventos del calendario (alimentan la Agenda del Inicio). Agregá filas (UID vacío). Participantes: nombres (de Equipo/Entidades) separados por « ; ».",
          [("UID", 16, None, True), ("Evento", 30, None, False), ("Fecha", 13, None, False),
           ("Inicio", 9, None, False), ("Fin", 9, None, False), ("Componente", 12, "comp", False),
           ("Modalidad", 13, "modal", False), ("Lugar", 22, None, False), ("Enlace conexión", 28, None, False),
           ("Participantes", 34, None, False), ("Formación", 11, "bool", False), ("URL documento", 28, None, False)],
          rows, extra=200)

# === 4. Equipo — Personas ===================================================
rows = [[p["uid"], p.get("apellido"), p.get("nombre"), ent_name.get(p.get("entidad_uid"), ""),
         p.get("rol"), p.get("componente"), p.get("telefono"), p.get("mail"), p.get("sexo")] for p in equipo]
add_sheet(wb, "Equipo - Personas",
          "Personas del equipo. Agregá filas (UID vacío). Entidad: nombre exacto de la pestaña « Equipo - Entidades ».",
          [("UID", 16, None, True), ("Apellido", 20, None, False), ("Nombre", 20, None, False),
           ("Entidad", 26, None, False), ("Rol", 22, None, False), ("Componente", 12, "comp", False),
           ("Teléfono", 16, None, False), ("Mail", 26, None, False), ("Sexo", 8, "sexo", False)],
          rows, extra=200)

# === 5. Equipo — Entidades ==================================================
rows = [[e["uid"], e.get("nombre")] for e in ents]
add_sheet(wb, "Equipo - Entidades",
          "Entidades (organismos, empresas…). Agregá filas (UID vacío). El nombre se referencia desde Personas y Eventos.",
          [("UID", 16, None, True), ("Entidad", 44, None, False)],
          rows, extra=120)

# === 6. Capacitaciones — Documentos =========================================
rows = [[c["uid"], c.get("subseccion"), c.get("componente"), c.get("titulo"), c.get("url"),
         b(c.get("confidencial")), b(c.get("publicar"))] for c in cap]
add_sheet(wb, "Capacitaciones - Documentos",
          "Material de formación, por subsección (EE / AyS / G). Agregá filas (UID vacío).",
          [("UID", 16, None, True), ("Subsección", 12, "subsec", False), ("Componente", 12, "comp", False),
           ("Título", 40, None, False), ("Enlace (URL)", 34, None, False),
           ("Confidencial", 12, "bool", False), ("Publicar", 12, "bool", False)],
          rows, extra=120)

# === 7. Subproyectos — Datos del edificio ===================================
rows = [[s["uid"], s.get("nombre"), s.get("tipologia"), s.get("seccion"), s.get("direccion"),
         s.get("lat"), s.get("lng"), s.get("superficie_m2"), s.get("notas")] for s in subs]
add_sheet(wb, "Subproyectos - Edificio",
          "Datos del edificio. Las 9 filas son fijas (no borrar). Solo se puede AGREGAR una escuela (fila al final, UID vacío, Sección=Escuelas).",
          [("UID", 16, None, True), ("Nombre", 32, None, False), ("Tipología", 11, "tip", False),
           ("Sección", 14, "secc", False), ("Dirección", 30, None, False), ("Latitud", 13, None, False),
           ("Longitud", 13, None, False), ("Superficie (m²)", 14, None, False), ("Notas", 36, None, False)],
          rows, extra=20)

# === 8. Subproyectos — Factibilidad (metricas) ==============================
def met_rows(escen, with_benef):
    out = []
    by = {(m["subproyecto_uid"], m["escenario"]): m for m in mets}
    for s in subs:
        m = by.get((s["uid"], escen), {})
        base = [s["uid"], sub_name.get(s["uid"], ""),
                m.get("demanda_kwh"), m.get("demanda_despues_kwh"), m.get("gei_antes_tco2"),
                m.get("gei_despues_tco2"), m.get("costo_ee_eur"), m.get("costo_otras_eur")]
        if with_benef:
            base += [m.get("benef_personal"), m.get("benef_personal_pct_muj"),
                     m.get("benef_usuarios"), m.get("benef_usuarios_pct_muj"),
                     m.get("benef_indirectos"), m.get("benef_indirectos_pct_muj")]
        out.append(base)
    return out

met_cols = [("subproyecto_uid", 16, None, True), ("Subproyecto", 30, None, True),
            ("Demanda antes (kWh)", 16, None, False), ("Demanda después (kWh)", 16, None, False),
            ("GEI antes (tCO₂)", 13, None, False), ("GEI después (tCO₂)", 13, None, False),
            ("Costo EE (€)", 13, None, False), ("Costo otras (€)", 14, None, False)]
benef_cols = [("Personal", 11, None, False), ("Personal % muj", 13, None, False),
              ("Usuarios", 11, None, False), ("Usuarios % muj", 13, None, False),
              ("Indirectos", 11, None, False), ("Indirectos % muj", 13, None, False)]
add_sheet(wb, "Subproyectos - Factibilidad",
          "Métricas del escenario de FACTIBILIDAD (+ beneficiarios). 9 filas fijas. Vacío = dato faltante (no 0).",
          met_cols + benef_cols, met_rows("faisabilidad", True))

# === 9. Subproyectos — Proyecto =============================================
add_sheet(wb, "Subproyectos - Proyecto",
          "Métricas del escenario de PROYECTO (sin beneficiarios). 9 filas fijas. Vacío = dato faltante (no 0).",
          met_cols, met_rows("proyecto", False))

# === 10. Subproyectos — Medidas =============================================
rows = [[m["subproyecto_uid"], sub_name.get(m["subproyecto_uid"], ""), m["medida"],
         MEDIDA_INFO.get(m["medida"], (m["medida"],))[0], b(m.get("activa")), m.get("texto"), m.get("kwh_anual")]
        for m in meds]
add_sheet(wb, "Subproyectos - Medidas",
          "Medidas por subproyecto (81 filas fijas, 9×9). Marcá Activa, escribí el detalle y, SOLO para medidas EE, el ahorro kWh/año.",
          [("subproyecto_uid", 16, None, True), ("Subproyecto", 28, None, True), ("medida", 16, None, True),
           ("Medida", 18, None, True), ("Activa", 9, "bool", False), ("Detalle", 40, None, False),
           ("kWh/año", 14, None, False)],
          rows, gray_kwh_for=lambda r: MEDIDA_INFO.get(r[2], ("", False))[1])

# === 11. Subproyectos — Gestión documentos ==================================
docs = [r for r in gest if r.get("tipo_linea") != "etapa"]
rows = [[r["uid"], r["subproyecto_uid"], sub_name.get(r["subproyecto_uid"], ""), r.get("titulo"),
         r.get("componente"), r.get("url"), r.get("estado"), r.get("fecha"),
         b(r.get("confidencial")), b(r.get("publicar"))] for r in docs]
add_sheet(wb, "Subproyectos - Gestión docs",
          "Documentos de gestión por subproyecto. Agregá filas (UID vacío) indicando subproyecto_uid (ver pestaña Edificio).",
          [("UID", 16, None, True), ("subproyecto_uid", 16, None, True), ("Subproyecto", 26, None, True),
           ("Título", 34, None, False), ("Componente", 12, "comp", False), ("Enlace (URL)", 30, None, False),
           ("Estado", 13, "estado", False), ("Fecha", 13, None, False),
           ("Confidencial", 12, "bool", False), ("Publicar", 12, "bool", False)],
          rows, extra=150)

# === 12. Subproyectos — Fases ===============================================
fases = [r for r in gest if r.get("tipo_linea") == "etapa"]
rows = [[r["subproyecto_uid"], sub_name.get(r["subproyecto_uid"], ""), r.get("fase"),
         FASE_NOMBRE.get(r.get("fase"), r.get("fase")), r.get("estado"),
         r.get("fecha_inicio"), r.get("fecha_fin")] for r in fases]
add_sheet(wb, "Subproyectos - Fases",
          "Fases por subproyecto (fijas). Completá Estado y fechas (inicio / fin).",
          [("subproyecto_uid", 16, None, True), ("Subproyecto", 26, None, True), ("fase", 18, None, True),
           ("Fase", 20, None, True), ("Estado", 14, "estado", False),
           ("Fecha inicio", 13, None, False), ("Fecha fin", 13, None, False)],
          rows)

wb.save(OUT)
print("OK ->", OUT)
for ws in wb.worksheets:
    print(f"  {ws.title}: {ws.max_row - 2 if ws.title != 'Instrucciones' else 0} filas de datos")
