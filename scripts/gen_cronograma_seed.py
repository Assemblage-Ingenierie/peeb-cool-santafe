# -*- coding: utf-8 -*-
"""Génère le SQL de planning (durées/dates/liaisons) des 9 sous-projets à partir
de l'Excel Hojaruta. `--report` = validation du mapping sans écrire ; sinon écrit
seed_plans.sql (durées + dates) et seed_enlaces.sql (liaisons) dans le scratchpad."""
import openpyxl, unicodedata, re, sys, os

XLSX = r"C:\Users\maelb\Downloads\APPLI - Hojaruta - cronograma (1).xlsx"
OUT = r"C:\Users\maelb\AppData\Local\Temp\claude\C--Users-maelb-github-PEEB-Santa-Fe\7c7ee731-0cdb-42e9-94fd-5f771d95a66a\scratchpad"

COLS = {4:'SUB-AIR',5:'SUB-ASV',6:'SUB-CENTENARIO',7:'SUB-CULLEN',8:'SUB-E67',9:'SUB-E1109',10:'SUB-E331',11:'SUB-E407',12:'SUB-E574'}

def norm(s):
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii','ignore').decode().lower()
    return re.sub(r'\s+',' ',s).strip().rstrip('.')

def pk(c): return '__fase__'+c

PHASE_ROW = {
 'estudios preliminares':('phase','estudios_preliminares'),
 'anteproyecto':('phase','anteproyecto'),
 'validacion de anteproyecto':('phase','validacion_anteproyecto'),
 'proyecto ejecutivo':('phase','proyecto_ejecutivo'),
 'redaccion de pliegos':('phase','redaccion_pliegos'),
 'no objecion afd sobre pliego':('phase','no_objecion_afd'),
 'licitacion':('phase','licitacion'),
 'publicacion del pliego':('card','gp-lic-publicacion'),
 'analisis y atribucion':('card','gp-lic-analisis'),
 'no objecion afd sobre resultado':('phase','no_objecion_afd_atribucion'),
 'negociacion y firma del contrato':('card','gp-lic-negociacion'),
 'no objecion afd sobre contrato':('phase','no_objecion_afd_contrato'),
 'obras de renovacion':('phase','obra'),
}

K_MEM='Redacción de memoria descriptiva del anteproyecto'
K_OTROS='Identificación de los otros planes de gestión relevantes para el proyecto'
K_PATR_ANT='Plan de gestión del Patrimonio - identificación de los requisitos'
K_AERO='Identificación de requisitos para intervenciones aeroportuarias'
K_HOSP='Identificación de requisitos para intervenciones hospitalarias'
K_ESC='Identificación de requisitos para intervenciones escolares'
K_PREL='Plan preliminar de continuidad de los servicios'
K_PRECAT='Pre-categorización provincial digital'
K_CAT='Categorización provincial'
K_CONT='Plan de continuidad de los servicios'
K_PATR_PE='Plan de gestión patrimonial'
K_L105='Lineamientos de los planes para la gestión de los aspectos ambientales'
K_L106='Lineamientos de los programas/planes para la gestión de trabajo, condiciones laborales y SST'
K_EXP='Participación de experto AyS en la redacción del pliego'
K_ASEG='Asegurar la integración de los lineamientos establecidos en la fase anterior'
K_VOF='Verificación de las ofertas AyS según criterios AyS'
K_VPL='Verificación de los Planes de Gestión AyS propuestos'
K_PGASC='Aprobación y seguimiento del PGASC'
K_CONF='Conformidad del cronograma con el plan de continuidad'
K_COORD='Coordinación y seguimiento de los planes solicitados'
K_RECL='Gestión de reclamos'

CARD_MAP = {
 ('estudios preliminares','elegibilidad y nivel de riesgo ficha de evaluacion (anexo 5)'):None,
 ('estudios preliminares','auditoria energetica'):None,
 ('estudios preliminares','diagnostico con perspectiva de genero'):None,
 ('estudios preliminares','actualizacion del modelo de simulacion'):'ee-ep-actualizacion-modelo',
 ('estudios preliminares','aprobacion del criterio peeb cool'):'ee-ep-aprobacion-criterio',
 ('estudios preliminares','formacion a los equipos de la ug / ministerio de linea sobre la incorporacion de la perspectiva de genero'):'genero-ep-formacion',
 ('anteproyecto','comprobacion del indicador peeb cool'):'ee-antep-comprobacion',
 ('anteproyecto','revision de proyecto con perspectiva de genero'):'genero-antep-revision',
 ('anteproyecto','validacion de las medidas con mujeres beneficiarias'):'genero-antep-validacion',
 ('anteproyecto','participacion de secretaria de mujeres, genero y diversidad'):'genero-antep-secretaria',
 ('anteproyecto','redaccion de memoria descriptiva del anteproyecto'):K_MEM,
 ('anteproyecto','plan de gestion del patrimonio - identificacion de los requisitos'):K_PATR_ANT,
 ('anteproyecto','identificacion de requisitos para intervenciones aeroportuarias'):K_AERO,
 ('anteproyecto','identificacion de requisitos para intervenciones hospitalarias'):K_HOSP,
 ('anteproyecto','identificacion de requisitos para intervenciones escolares'):K_ESC,
 ('anteproyecto','identificacion de los otros planes de gestion relevantes para el proyecto'):K_OTROS,
 ('anteproyecto','plan preliminar de continuidad de los servicios'):K_PREL,
 ('proyecto ejecutivo','comprobacion del indicador peeb cool'):'ee-pe-comprobacion',
 ('proyecto ejecutivo','revision de especificaciones tecnicas de los materiales y equipos relativos a la ee'):'ee-pe-especificaciones',
 ('proyecto ejecutivo','revision de proyecto con perspectiva de genero'):'genero-pe-revision',
 ('proyecto ejecutivo','pre-categorizacion provincial digital'):K_PRECAT,
 ('proyecto ejecutivo','categorizacion provincial'):K_CAT,
 ('proyecto ejecutivo','plan de continuidad de los servicios'):K_CONT,
 ('proyecto ejecutivo','plan de gestion patrimonial'):K_PATR_PE,
 ('proyecto ejecutivo','lineamientos de los planes para la gestion de los aspectos ambientales'):K_L105,
 ('proyecto ejecutivo','lineamientos de los programas/planes para la gestion de trabajo, condiciones laboralesy sst'):K_L106,
 ('redaccion de pliegos','participacion de experto ays en la redaccion del pliego'):K_EXP,
 ('redaccion de pliegos','asegurar la integracion de los lineamientos establecidos en la fase anterior'):K_ASEG,
 ('redaccion de pliegos','revision y correccion del lenguaje en documentos de licitacion'):'genero-pliegos-lenguaje',
 ('redaccion de pliegos','incorporar clausulas que valoren a oferentes con politicas de genero'):'genero-pliegos-clausulas',
 ('redaccion de pliegos','inclusion de politicas de igualdad y de genero en los pliegos de condiciones de las licitaciones'):'genero-pliegos-inclusion',
 ('redaccion de pliegos','definicion de criterios de evaluacion especificos para valorar oferentes con politicas de genero'):'genero-pliegos-criterios',
 ('redaccion de pliegos','revision de documentos de licitacion con respecto a empresas lideradas por mujeres (elm)'):'genero-pliegos-elm',
 ('analisis y atribucion','evaluacion de ofertas con perspectiva de genero'):'genero-licitacion-evaluacion',
 ('analisis y atribucion','verificacion de las ofertas ays segun criterios ays'):K_VOF,
 ('analisis y atribucion','verificacion de los planes de gestion ays propuestos'):K_VPL,
 ('obras de renovacion','aprobacion y seguimiento del pgasc'):K_PGASC,
 ('obras de renovacion','conformidad del cronograma con el plan de continuidad'):K_CONF,
 ('obras de renovacion','coordinacion y seguimiento de los planes solicitados'):K_COORD,
 ('obras de renovacion','gestion de reclamos'):K_RECL,
}

REF_MAP = {
 'actualizacion del modelo':('card','ee-ep-actualizacion-modelo'),
 'anteproyecto':('phase','anteproyecto'),
 'ante proyecto':('phase','anteproyecto'),
 'proyecto':('phase','proyecto_ejecutivo'),
 'proyecto ejecutivo':('phase','proyecto_ejecutivo'),
 'valiidacion de ante proyecto':('phase','validacion_anteproyecto'),
 'redaccion de pliegos':('phase','redaccion_pliegos'),
 'no objecion afd pliegos':('phase','no_objecion_afd'),
 'publicacion de pliego':('card','gp-lic-publicacion'),
 'analisis y atribucion':('card','gp-lic-analisis'),
 'obras de renovacion':('phase','obra'),
 'no objecion afd sobre resultado':('phase','no_objecion_afd_atribucion'),
 'negociacion y firma del contrato':('card','gp-lic-negociacion'),
 'no objecion afd sobre contrato':('phase','no_objecion_afd_contrato'),
 'redaccion de memoria descriptiva':('card',K_MEM),
 'precategorizacion provincial digital':('card',K_PRECAT),
}

# Corrections d'année (l'Excel date certaines fases en 2026 par erreur → 2027) :
#  - anteproyecto : mars 2026 → 2027 (confirmé client)
#  - estudios preliminares des écoles : avril 2026 → 2027 (confirmé client)
def fixyear(d):
    if d == '2026-03-01': return '2027-03-01'
    if d == '2026-04-01': return '2027-04-01'
    return d

def unit(u):
    u=norm(u)
    if u.startswith('mes') or u.startswith('mois'): return 'mes'
    if u.startswith('seman') or u.startswith('semain'): return 'semana'
    if u.startswith('dia'): return 'dia'
    return None

def parse_dur(d):
    m=re.match(r'(\d+)\s+(\S+)', d or '')
    return (int(m.group(1)), unit(m.group(2))) if m else (None,None)

def target_key(tgt):
    kind,code = tgt
    return pk(code) if kind=='phase' else code

def parse_start(s, errors):
    sn=norm(s)
    m=re.match(r'(\d+)\s+(\S+)\s+antes del (inicio|fin) del? (.+)$', sn)
    if m:
        tgt=REF_MAP.get(m.group(4))
        if not tgt: errors.append('REF? '+repr(m.group(4))+' <- '+s); return None
        return ('dep', target_key(tgt), m.group(3), -int(m.group(1)), unit(m.group(2)))
    m=re.match(r'(inicio|fin) del? (.+)$', sn)
    if m:
        tgt=REF_MAP.get(m.group(2))
        if not tgt: errors.append('REF? '+repr(m.group(2))+' <- '+s); return None
        return ('dep', target_key(tgt), m.group(1), 0, 'dia')
    errors.append('START? '+repr(s)); return None

ws=openpyxl.load_workbook(XLSX,data_only=True)["Feuille 1"]
def fr(c):
    f=c.fill; return (getattr(f.fgColor,'rgb',None) if f and f.patternType else '') or ''
def cv(c):
    x=c.value
    if x in (None,''): return ('empty',None)
    if hasattr(x,'strftime'): return ('date', x.strftime('%Y-%m-%d'))
    return ('str', str(x).strip())

errors=[]; unmapped=set()
data={u:[] for u in COLS.values()}
section=None; r=2
while r<=107:
    lbl = cv(ws.cell(r,3))[1]
    if lbl and norm(lbl)=='fecha inicio':
        nn=norm(ws.cell(r,2).value)
        if nn in PHASE_ROW:
            section=nn; target=PHASE_ROW[nn]
        else:
            target=CARD_MAP.get((section,nn),'__MISSING__')
            if target=='__MISSING__': unmapped.add((section,nn))
        tnorm = ('card',target) if isinstance(target,str) else target
        for col,uid in COLS.items():
            if fr(ws.cell(r,col))=='FF000000': continue
            if target in (None,'__MISSING__'): continue
            kind,_=tnorm; key=target_key(tnorm)
            dv,du=parse_dur(cv(ws.cell(r+1,col))[1])
            typ,val=cv(ws.cell(r,col))
            st=('abs',val) if typ=='date' else (parse_start(val,errors) if typ=='str' else None)
            data[uid].append((kind,key,dv,du,st))
        r+=2
    else: r+=1

def q(s): return str(s).replace("'","''")
plans=[]; enl=[]
for uid,rows in data.items():
    enl.append(f"delete from public.peebcoolsf_roadmap_enlace where feuille='{uid}';")
    for kind,key,dv,du,st in rows:
        fi = f"'{fixyear(st[1])}'" if (st and st[0]=='abs') else 'null'
        dvs = str(dv) if dv is not None else 'null'
        dus = f"'{du}'" if du else 'null'
        if kind=='phase':
            code=key.replace('__fase__','')
            plans.append(f"update public.peebcoolsf_gestion_lineas set dur_valor={dvs}, dur_unidad={dus}, fecha_inicio={fi} where subproyecto_uid='{uid}' and tipo_linea='etapa' and fase='{code}';")
        else:
            plans.append(f"insert into public.peebcoolsf_roadmap_estado (feuille,tarea_key,dur_valor,dur_unidad,fecha_inicio) values ('{uid}','{q(key)}',{dvs},{dus},{fi}) on conflict (feuille,tarea_key) do update set dur_valor=excluded.dur_valor, dur_unidad=excluded.dur_unidad, fecha_inicio=excluded.fecha_inicio;")
        if st and st[0]=='dep':
            tk,punto,off,ou=st[1:]
            enl.append(f"insert into public.peebcoolsf_roadmap_enlace (feuille,desde,hacia,punto,desfase_valor,desfase_unidad) values ('{uid}','{q(tk)}','{q(key)}','{punto}',{off},'{ou}') on conflict (feuille,desde,hacia) do update set punto=excluded.punto, desfase_valor=excluded.desfase_valor, desfase_unidad=excluded.desfase_unidad;")

print("=== UNMAPPED ROWS ==="); [print(x) for x in sorted(unmapped)]
print("=== ERRORS ==="); [print(x) for x in errors[:50]]
print("=== COUNTS per sub ===")
for uid,rows in data.items():
    ph=sum(1 for k,*_ in rows if k=='phase'); cd=sum(1 for k,*_ in rows if k=='card')
    ab=sum(1 for *_,st in rows if st and st[0]=='abs'); dp=sum(1 for *_,st in rows if st and st[0]=='dep')
    print(f"  {uid:16} phases={ph} cards={cd} abs={ab} liaisons={dp}")
print("TOTAL plans:",len(plans)," enlace stmts:",len(enl))

if '--report' not in sys.argv:
    os.makedirs(OUT,exist_ok=True)
    open(os.path.join(OUT,'seed_plans.sql'),'w',encoding='utf-8').write("\n".join(plans))
    open(os.path.join(OUT,'seed_enlaces.sql'),'w',encoding='utf-8').write("\n".join(enl))
    # --- Compact (multi-rows) pour les sous-projets != SUB-AIR (AIR déjà seedé) ---
    NONAIR=[u for u in COLS.values() if u!='SUB-AIR']
    gest=[]; estv=[]; enlv=[]
    for uid in NONAIR:
        for kind,key,dv,du,st in data[uid]:
            dvs=str(dv) if dv is not None else 'null'; dus=f"'{du}'" if du else 'null'
            fi=f"'{fixyear(st[1])}'" if (st and st[0]=='abs') else 'null'
            if kind=='phase':
                gest.append(f"update public.peebcoolsf_gestion_lineas set dur_valor={dvs}, dur_unidad={dus}, fecha_inicio={fi} where subproyecto_uid='{uid}' and tipo_linea='etapa' and fase='{key.replace('__fase__','')}';")
            else:
                estv.append(f"('{uid}','{q(key)}',{dvs},{dus},{fi})")
            if st and st[0]=='dep':
                tk,punto,off,ou=st[1:]
                enlv.append(f"('{uid}','{q(tk)}','{q(key)}','{punto}',{off},'{ou}')")
    inlist=",".join("'"+u+"'" for u in NONAIR)
    est_sql="insert into public.peebcoolsf_roadmap_estado (feuille,tarea_key,dur_valor,dur_unidad,fecha_inicio) values\n"+",\n".join(estv)+"\non conflict (feuille,tarea_key) do update set dur_valor=excluded.dur_valor, dur_unidad=excluded.dur_unidad, fecha_inicio=excluded.fecha_inicio;"
    ENLHEAD="insert into public.peebcoolsf_roadmap_enlace (feuille,desde,hacia,punto,desfase_valor,desfase_unidad) values\n"
    ENLTAIL="\non conflict (feuille,desde,hacia) do update set punto=excluded.punto, desfase_valor=excluded.desfase_valor, desfase_unidad=excluded.desfase_unidad;"
    half=len(enlv)//2
    enl1=f"delete from public.peebcoolsf_roadmap_enlace where feuille in ({inlist});\n"+ENLHEAD+",\n".join(enlv[:half])+ENLTAIL
    enl2=ENLHEAD+",\n".join(enlv[half:])+ENLTAIL
    open(os.path.join(OUT,'compact_gest.sql'),'w',encoding='utf-8').write("\n".join(gest))
    open(os.path.join(OUT,'compact_est.sql'),'w',encoding='utf-8').write(est_sql)
    open(os.path.join(OUT,'compact_enl1.sql'),'w',encoding='utf-8').write(enl1)
    open(os.path.join(OUT,'compact_enl2.sql'),'w',encoding='utf-8').write(enl2)
    print(f"COMPACT non-AIR: gest={len(gest)} lignes, est={len(estv)} tuples, enl={len(enlv)} tuples")
    print("sizes (KB):", round(len('\n'.join(gest))/1024,1), round(len(est_sql)/1024,1), round(len(enl_sql)/1024,1))
