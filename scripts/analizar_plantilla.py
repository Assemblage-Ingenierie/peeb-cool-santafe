# -*- coding: utf-8 -*-
"""
Analyse la plantilla remplie vs le modèle d'origine et affiche un résumé des
changements (modifs / nouvelles lignes / avertissements). NE TOUCHE PAS la base.
"""
import os, sys, json, ssl, urllib.request, datetime, io
import openpyxl

# Force UTF-8 stdout (console Windows)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # racine du repo (depuis scripts/)
ORIG = os.path.join(HERE, "PEEB - Plantilla datos Admin.xlsx")
FILL = os.path.join(HERE, "PEEB - Plantilla datos Admin - remplie.xlsx")

# --- maps noms -> uid depuis la base (entidades / equipo) ------------------
def load_env(path):
    env = {}
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1); env[k.strip()] = v.strip().strip('"').strip("'")
    return env
env = load_env(os.path.join(HERE, ".env.local"))
URL = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/"); KEY = env["SUPABASE_SERVICE_ROLE_KEY"]
def fetch(t, p="select=*"):
    req = urllib.request.Request(f"{URL}/rest/v1/{t}?{p}", headers={"apikey": KEY, "Authorization": f"Bearer {KEY}"})
    with urllib.request.urlopen(req, context=ssl.create_default_context()) as r:
        return json.loads(r.read().decode())
ents = fetch("peebcoolsf_entidades"); equipo = fetch("peebcoolsf_equipo")
def plabel(p): return (f'{(p.get("apellido") or "").strip()}, {(p.get("nombre") or "").strip()}').strip().strip(",").strip()
ent_by_name = {e["nombre"].strip().lower(): e["uid"] for e in ents}
part_by_name = {**{plabel(p).lower(): p["uid"] for p in equipo}, **ent_by_name}

# --- normalisation ----------------------------------------------------------
def n_text(v):
    if v is None: return None
    s = str(v).strip(); return s or None
def n_bool(v):
    return str(v).strip().lower() in ("sí","si","x","true","verdadero","1","yes","oui") if v is not None else False
def n_num(v, integer=False):
    if v is None or (isinstance(v, str) and not v.strip()): return None
    try: f = float(str(v).replace(",", ".")) if isinstance(v, str) else float(v)
    except Exception: return ("ERR", v)
    return int(round(f)) if integer else round(f, 6)
def n_date(v):
    if v is None or (isinstance(v, str) and not v.strip()): return None
    if isinstance(v, (datetime.datetime, datetime.date)): return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    try: datetime.datetime.strptime(s, "%Y-%m-%d"); return s
    except Exception: return ("ERR", s)
def n_time(v):
    if v is None or (isinstance(v, str) and not v.strip()): return None
    if isinstance(v, (datetime.datetime, datetime.time)): return v.strftime("%H:%M")
    s = str(v).strip(); return s[:5]

# colonnes par feuille : (db_field|None, kind). None = colonne de référence (ignorée).
T, B, I, F, D, TM = "text", "bool", "int", "float", "date", "time"
SHEETS = [
 ("GP - Documentación", "documentacion_gp", ["uid"], "append", None, None,
  [("uid",T),("nombre_documento",T),("componente",T),("url",T),("confidencial",B),("publicar",B)]),
 ("GP - Gestión financiera", "gestion_financiera", ["uid"], "append", None, None,
  [("uid",T),("titulo",T),("url",T),("confidencial",B),("publicar",B)]),
 ("Calendario - Eventos", "eventos", ["uid"], "append", None, None,
  [("uid",T),("nombre",T),("fecha",D),("hora_inicio",TM),("hora_fin",TM),("componente",T),
   ("modalidad",T),("lugar",T),("url_conexion",T),("participantes","parts"),("formacion",B),("url_documento",T)]),
 ("Equipo - Personas", "equipo", ["uid"], "append", None, None,
  [("uid",T),("apellido",T),("nombre",T),("entidad_uid","ent"),("rol",T),("componente",T),("telefono",T),("mail",T),("sexo",T)]),
 ("Equipo - Entidades", "entidades", ["uid"], "append", None, None,
  [("uid",T),("nombre",T)]),
 ("Capacitaciones - Documentos", "capacitaciones_documentos", ["uid"], "append", None, None,
  [("uid",T),("subseccion",T),("componente",T),("titulo",T),("url",T),("confidencial",B),("publicar",B)]),
 ("Subproyectos - Edificio", "subproyectos", ["uid"], "fixed", None, None,
  [("uid",T),("nombre",T),("tipologia",T),("seccion",T),("direccion",T),("lat",F),("lng",F),("superficie_m2",F),("notas",T)]),
 ("Subproyectos - Factibilidad", "metricas", ["subproyecto_uid"], "fixed", "faisabilidad", None,
  [("subproyecto_uid",T),(None,T),("demanda_kwh",F),("demanda_despues_kwh",F),("gei_antes_tco2",F),("gei_despues_tco2",F),
   ("costo_ee_eur",F),("costo_otras_eur",F),("benef_personal",I),("benef_personal_pct_muj",F),
   ("benef_usuarios",I),("benef_usuarios_pct_muj",F),("benef_indirectos",I),("benef_indirectos_pct_muj",F)]),
 ("Subproyectos - Proyecto", "metricas", ["subproyecto_uid"], "fixed", "proyecto", None,
  [("subproyecto_uid",T),(None,T),("demanda_kwh",F),("demanda_despues_kwh",F),("gei_antes_tco2",F),("gei_despues_tco2",F),
   ("costo_ee_eur",F),("costo_otras_eur",F)]),
 ("Subproyectos - Medidas", "medidas", ["subproyecto_uid","medida"], "fixed", None, None,
  [("subproyecto_uid",T),(None,T),("medida",T),(None,T),("activa",B),("texto",T),("kwh_anual",F)]),
 ("Subproyectos - Gestión docs", "gestion_lineas", ["uid"], "append", None, "documento",
  [("uid",T),("subproyecto_uid",T),(None,T),("titulo",T),("componente",T),("url",T),("estado",T),("fecha",D),("confidencial",B),("publicar",B)]),
 ("Subproyectos - Fases", "gestion_lineas", ["subproyecto_uid","fase"], "fixed", None, "etapa",
  [("subproyecto_uid",T),(None,T),("fase",T),(None,T),("estado",T),("fecha_inicio",D),("fecha_fin",D)]),
]

def norm_val(kind, v, warns, ctx):
    if kind == B: return n_bool(v)
    if kind == I: r = n_num(v, True)
    elif kind == F: r = n_num(v)
    elif kind == D: r = n_date(v)
    elif kind == TM: return n_time(v)
    elif kind == "ent":
        t = n_text(v)
        if t is None: return None
        uid = ent_by_name.get(t.lower())
        if not uid: warns.append(f"{ctx}: entidad desconocida « {t} »")
        return uid or ("ERR", t)
    elif kind == "parts":
        t = n_text(v)
        if t is None: return []
        out = []
        for nm in [x.strip() for x in t.replace("\n",";").split(";") if x.strip()]:
            uid = part_by_name.get(nm.lower())
            if not uid: warns.append(f"{ctx}: participante desconocido « {nm} »")
            out.append(uid or f"?{nm}")
        return out
    else: return n_text(v)
    if isinstance(r, tuple): warns.append(f"{ctx}: valor numérico/fecha inválido « {r[1]} »"); return None
    return r

def read_sheet(wb, cfg):
    name, table, keyf, mode, escen, tipo, cols = cfg
    ws = wb[name]; recs = {}; warns = []; order = []
    for i in range(3, ws.max_row + 1):
        raw = [ws.cell(i, j + 1).value for j in range(len(cols))]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw):
            continue
        rec = {}; ctx = f"{name} fila {i}"
        for j, (fld, kind) in enumerate(cols):
            if fld is None: continue
            rec[fld] = norm_val(kind, raw[j], warns, ctx)
        # clé
        k = tuple(rec.get(f) for f in keyf)
        recs_key = k if any(k) else ("__NEW__", i)
        rec["__ctx__"] = ctx
        recs[recs_key] = rec; order.append(recs_key)
    return recs, warns

wb_o = openpyxl.load_workbook(ORIG, data_only=True)
wb_f = openpyxl.load_workbook(FILL, data_only=True)

EDITABLE_SKIP = {"__ctx__"}
grand_ins = grand_upd = 0
print("=" * 70)
print("RESUMEN DE CAMBIOS  (modelo original  ->  archivo remplie)")
print("=" * 70)
all_warns = []
for cfg in SHEETS:
    name = cfg[0]; keyf = cfg[2]; mode = cfg[3]
    o, _ = read_sheet(wb_o, cfg)
    f, warns = read_sheet(wb_f, cfg)
    all_warns += warns
    inserts = []; updates = []
    for k, rec in f.items():
        is_new = isinstance(k, tuple) and k and k[0] == "__NEW__"
        # ligne « append » sans uid (uid vide) = nouvelle
        if is_new or (mode == "append" and not rec.get(keyf[0])):
            if any(v not in (None, [], False) for kk, v in rec.items() if kk not in EDITABLE_SKIP and kk not in keyf):
                inserts.append(rec)
            continue
        base = o.get(k)
        if base is None:
            updates.append((rec, {"(nueva clave no existente en original)": (None, k)})); continue
        diffs = {}
        for fld, val in rec.items():
            if fld in EDITABLE_SKIP: continue
            bval = base.get(fld)
            if fld in keyf: continue
            if (val if val != [] else None) != (bval if bval != [] else None):
                diffs[fld] = (bval, val)
        if diffs: updates.append((rec, diffs))
    # rapport feuille
    if inserts or updates or warns:
        print(f"\n## {name}  ->  {len(updates)} modif · {len(inserts)} nuevas" + (f" · {len(warns)} avisos" if warns else ""))
        for rec, diffs in updates[:60]:
            key = " / ".join(str(rec.get(x)) for x in keyf)
            ch = ", ".join(f"{fld}: {b!r}→{v!r}" for fld, (b, v) in diffs.items())
            print(f"   ~ {key}: {ch}")
        if len(updates) > 60: print(f"   … (+{len(updates)-60} modifs)")
        for rec in inserts[:60]:
            disp = {kk: vv for kk, vv in rec.items() if kk not in EDITABLE_SKIP and vv not in (None, [], False)}
            print(f"   + NUEVA: {disp}")
        if len(inserts) > 60: print(f"   … (+{len(inserts)-60} nuevas)")
    grand_ins += len(inserts); grand_upd += len(updates)

print("\n" + "=" * 70)
print(f"TOTAL: {grand_upd} modificaciones · {grand_ins} nuevas filas")
if all_warns:
    print(f"\n!!! AVISOS ({len(all_warns)}):")
    for w in all_warns: print("   -", w)
else:
    print("Sin avisos.")
