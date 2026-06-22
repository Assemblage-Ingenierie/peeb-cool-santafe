# -*- coding: utf-8 -*-
"""
Prépare l'import : compare le fichier rempli à la BASE ACTUELLE (baseline = DB),
classe en UPDATE / INSERT, applique « apellidos en MAJUSCULES », affiche un résumé
des NOUVEAUTÉS et écrit import_generado.sql (idempotent). N'EXÉCUTE RIEN.
"""
import os, sys, json, ssl, urllib.request, datetime, io
import openpyxl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # racine du repo (depuis scripts/)
FILL = os.path.join(HERE, "PEEB - Plantilla datos Admin - remplie.xlsx")
SQL_OUT = os.path.join(HERE, "import_generado.sql")

def load_env(p):
    e = {}
    for line in open(p, encoding="utf-8"):
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1); e[k.strip()] = v.strip().strip('"').strip("'")
    return e
env = load_env(os.path.join(HERE, ".env.local"))
URL = env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/"); KEY = env["SUPABASE_SERVICE_ROLE_KEY"]
def fetch(t, p="select=*"):
    req = urllib.request.Request(f"{URL}/rest/v1/{t}?{p}", headers={"apikey": KEY, "Authorization": f"Bearer {KEY}"})
    with urllib.request.urlopen(req, context=ssl.create_default_context()) as r:
        return json.loads(r.read().decode())

db = {t: fetch("peebcoolsf_" + t) for t in
      ["documentacion_gp","gestion_financiera","eventos","equipo","entidades",
       "capacitaciones_documentos","subproyectos","metricas","medidas","gestion_lineas"]}
db_uids = {t: {r.get("uid") for r in rows if "uid" in r} for t, rows in db.items()}

def n_text(v, fld=None):
    if v is None: return None
    s = str(v).strip()
    if not s: return None
    if fld == "apellido": s = s.upper()  # règle : apellidos en MAJUSCULES
    return s
def n_bool(v):
    return str(v).strip().lower() in ("sí","si","x","true","verdadero","1","yes","oui") if v is not None else False
def n_num(v, integer=False):
    if v is None or (isinstance(v, str) and not v.strip()): return None
    try: f = float(str(v).replace(",", ".")) if isinstance(v, str) else float(v)
    except Exception: return ("ERR", v)
    return int(round(f)) if integer else round(f, 6)
def n_date(v):
    if v is None or (isinstance(v, str) and not v.strip()): return None
    if isinstance(v, (datetime.datetime, datetime.date)): return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    try: datetime.datetime.strptime(s[:10], "%Y-%m-%d"); return s[:10]
    except Exception: return ("ERR", s)
def n_time(v):
    if v is None or (isinstance(v, str) and not v.strip()): return None
    if isinstance(v, (datetime.datetime, datetime.time)): return v.strftime("%H:%M")
    return str(v).strip()[:5]

T, B, I, F, D, TM = "text", "bool", "int", "float", "date", "time"
SHEETS = [
 ("GP - Documentación","documentacion_gp",["uid"],None,None,
  [("uid",T),("nombre_documento",T),("componente",T),("url",T),("confidencial",B),("publicar",B)]),
 ("GP - Gestión financiera","gestion_financiera",["uid"],None,None,
  [("uid",T),("titulo",T),("url",T),("confidencial",B),("publicar",B)]),
 ("Calendario - Eventos","eventos",["uid"],None,None,
  [("uid",T),("nombre",T),("fecha",D),("hora_inicio",TM),("hora_fin",TM),("componente",T),
   ("modalidad",T),("lugar",T),("url_conexion",T),("participantes","parts"),("formacion",B),("url_documento",T)]),
 ("Equipo - Personas","equipo",["uid"],None,None,
  [("uid",T),("apellido",T),("nombre",T),("entidad_uid","ent"),("rol",T),("componente",T),("telefono",T),("mail",T),("sexo",T)]),
 ("Equipo - Entidades","entidades",["uid"],None,None,[("uid",T),("nombre",T)]),
 ("Capacitaciones - Documentos","capacitaciones_documentos",["uid"],None,None,
  [("uid",T),("subseccion",T),("componente",T),("titulo",T),("url",T),("confidencial",B),("publicar",B)]),
 ("Subproyectos - Edificio","subproyectos",["uid"],None,None,
  [("uid",T),("nombre",T),("tipologia",T),("seccion",T),("direccion",T),("lat",F),("lng",F),("superficie_m2",F),("notas",T)]),
 ("Subproyectos - Factibilidad","metricas",["subproyecto_uid"],"faisabilidad",None,
  [("subproyecto_uid",T),(None,T),("demanda_kwh",F),("demanda_despues_kwh",F),("gei_antes_tco2",F),("gei_despues_tco2",F),
   ("costo_ee_eur",F),("costo_otras_eur",F),("benef_personal",I),("benef_personal_pct_muj",F),
   ("benef_usuarios",I),("benef_usuarios_pct_muj",F),("benef_indirectos",I),("benef_indirectos_pct_muj",F)]),
 ("Subproyectos - Proyecto","metricas",["subproyecto_uid"],"proyecto",None,
  [("subproyecto_uid",T),(None,T),("demanda_kwh",F),("demanda_despues_kwh",F),("gei_antes_tco2",F),("gei_despues_tco2",F),
   ("costo_ee_eur",F),("costo_otras_eur",F)]),
 ("Subproyectos - Medidas","medidas",["subproyecto_uid","medida"],None,None,
  [("subproyecto_uid",T),(None,T),("medida",T),(None,T),("activa",B),("texto",T),("kwh_anual",F)]),
 ("Subproyectos - Gestión docs","gestion_lineas",["uid"],None,"documento",
  [("uid",T),("subproyecto_uid",T),(None,T),("titulo",T),("componente",T),("url",T),("estado",T),("fecha",D),("confidencial",B),("publicar",B)]),
 ("Subproyectos - Fases","gestion_lineas",["subproyecto_uid","fase"],None,"etapa",
  [("subproyecto_uid",T),(None,T),("fase",T),(None,T),("estado",T),("fecha_inicio",D),("fecha_fin",D)]),
]

# maps noms -> uid : base + nouvelles entidades de la feuille remplie
def plabel(p): return (f'{(p.get("apellido") or "").strip()}, {(p.get("nombre") or "").strip()}').strip().strip(",").strip()
ent_by_name = {e["nombre"].strip().lower(): e["uid"] for e in db["entidades"]}
wb_f = openpyxl.load_workbook(FILL, data_only=True)
ws_e = wb_f["Equipo - Entidades"]
for i in range(3, ws_e.max_row + 1):
    uid = n_text(ws_e.cell(i,1).value); nom = n_text(ws_e.cell(i,2).value)
    if nom: ent_by_name[nom.lower()] = uid or "(auto)"
part_by_name = {**{plabel(p).lower(): p["uid"] for p in db["equipo"]}, **ent_by_name}

ENUMS = {"componente":{"GP","EE","AyS","G"}, "tipologia":{"A","H","E"},
         "seccion":{"Aeropuertos","Hospitales","Escuelas"}, "estado":{"en_proceso","terminado"},
         "modalidad":{"Presencial","Virtual"}, "sexo":{"F","M","X"}, "subseccion":{"EE","AyS","G"}}
warns = []
def fill_val(kind, v, ctx, fld):
    if kind == B: return n_bool(v)
    if kind == I:
        r = n_num(v, True)
    elif kind == F:
        r = n_num(v)
    elif kind == D:
        r = n_date(v)
    elif kind == TM:
        return n_time(v)
    elif kind == "ent":
        t = n_text(v)
        if not t: return None
        u = ent_by_name.get(t.lower())
        if not u: warns.append(f"{ctx}: entidad desconocida « {t} »")
        return u
    elif kind == "parts":
        t = n_text(v)
        if not t: return []
        out = []
        for nm in [x.strip() for x in t.replace("\n",";").split(";") if x.strip()]:
            u = part_by_name.get(nm.lower())
            if not u: warns.append(f"{ctx}: participante desconocido « {nm} »")
            out.append(u or nm)
        return out
    else:
        r = n_text(v, fld)
        if fld in ENUMS and r is not None and r not in ENUMS[fld]:
            warns.append(f"{ctx}: valor inválido « {r} » para {fld}")
        return r
    if isinstance(r, tuple): warns.append(f"{ctx}: valor inválido « {r[1]} » ({fld})"); return None
    return r

def db_val(kind, r, fld):
    v = r.get(fld)
    if kind == B: return bool(v)
    if kind == I: return int(v) if v is not None else None
    if kind == F: return round(float(v), 6) if v is not None else None
    if kind == D: return str(v)[:10] if v else None
    if kind == TM: return str(v)[:5] if v else None
    if kind == "ent": return r.get("entidad_uid") or None
    if kind == "parts": return list(r.get("participantes") or [])
    return n_text(v, fld)

def read_fill(cfg):
    name, table, keyf, escen, tipo, cols = cfg
    ws = wb_f[name]; out = []
    for i in range(3, ws.max_row + 1):
        raw = [ws.cell(i, j+1).value for j in range(len(cols))]
        if all(v is None or (isinstance(v,str) and not v.strip()) for v in raw): continue
        rec, ctx = {}, f"{name} f{i}"
        for j,(fld,kind) in enumerate(cols):
            if fld is None: continue
            rec[fld] = fill_val(kind, raw[j], ctx, fld)
        rec["__ctx__"] = ctx
        out.append(rec)
    return out

def db_index(cfg):
    name, table, keyf, escen, tipo, cols = cfg
    idx = {}
    for r in db[table]:
        if escen and r.get("escenario") != escen: continue
        if tipo == "etapa" and r.get("tipo_linea") != "etapa": continue
        if tipo == "documento" and r.get("tipo_linea") == "etapa": continue
        rec = {fld: db_val(kind, r, fld) for fld, kind in cols if fld is not None}
        k = tuple(rec.get(x) for x in keyf)
        idx[k] = rec
    return idx

def sqlv(v):
    if v is None: return "NULL"
    if isinstance(v, bool): return "true" if v else "false"
    if isinstance(v, (int, float)): return repr(v)
    if isinstance(v, list):  # colonne text[] (participantes)
        if not v: return "ARRAY[]::text[]"
        return "ARRAY[" + ",".join("'" + str(x).replace("'", "''") + "'" for x in v) + "]::text[]"
    return "'" + str(v).replace("'", "''") + "'"

sql, n_upd, n_ins = [], 0, 0
order = ["Equipo - Entidades","Equipo - Personas","GP - Documentación","GP - Gestión financiera",
         "Capacitaciones - Documentos","Calendario - Eventos","Subproyectos - Edificio",
         "Subproyectos - Factibilidad","Subproyectos - Proyecto","Subproyectos - Medidas",
         "Subproyectos - Gestión docs","Subproyectos - Fases"]
by_name = {c[0]: c for c in SHEETS}

print("=" * 70); print("PLAN DE IMPORTACIÓN — novedades vs base actual (no se escribe nada)"); print("=" * 70)
for nm in order:
    cfg = by_name[nm]; name, table, keyf, escen, tipo, cols = cfg
    frows = read_fill(cfg); base = db_index(cfg)
    ups, ins = [], []
    for r in frows:
        key = tuple(r.get(x) for x in keyf)
        uid_keyed = "uid" in keyf
        if uid_keyed and r.get("uid") not in db_uids[table]:
            ins.append(r); continue
        b = base.get(key)
        if b is None:
            ins.append(r); continue
        diff = {fld: r.get(fld) for fld, kind in cols
                if fld is not None and fld not in keyf
                and (r.get(fld) if r.get(fld) != [] else None) != (b.get(fld) if b.get(fld) != [] else None)}
        if diff: ups.append((r, key, diff))
    for r, key, diff in ups:
        setp = ", ".join(f"{k}={sqlv(v)}" for k, v in diff.items())
        wherep = " AND ".join(f"{k}={sqlv(r.get(k))}" for k in keyf)
        if escen: wherep += f" AND escenario={sqlv(escen)}"
        if tipo: wherep += f" AND tipo_linea={sqlv(tipo)}"
        sql.append(f"UPDATE public.peebcoolsf_{table} SET {setp} WHERE {wherep};")
    for r in ins:
        fields = {k: v for k, v in r.items() if k != "__ctx__" and v not in (None, [])}
        if escen: fields["escenario"] = escen
        if tipo: fields["tipo_linea"] = tipo
        cols_ = ", ".join(fields.keys()); vals_ = ", ".join(sqlv(v) for v in fields.values())
        sql.append(f"INSERT INTO public.peebcoolsf_{table} ({cols_}) SELECT {vals_} "
                   f"WHERE NOT EXISTS (SELECT 1 FROM public.peebcoolsf_{table} WHERE uid={sqlv(r.get('uid'))});")
    n_upd += len(ups); n_ins += len(ins)
    if ups or ins:
        print(f"\n## {name}: {len(ups)} updates · {len(ins)} inserts")
        for r, key, diff in ups[:50]:
            ch = ", ".join(f"{k}={v!r}" for k, v in diff.items())
            print(f"   ~ {'/'.join(str(x) for x in key)}: {ch[:240]}")
        for r in ins[:50]:
            show = {k: v for k, v in r.items() if k != "__ctx__" and v not in (None, [], False)}
            print(f"   + {show}")

# Filet de sécurité : normaliser tout apellido existant en MAJUSCULES.
sql.append("UPDATE public.peebcoolsf_equipo SET apellido = upper(apellido) WHERE apellido <> upper(apellido);")

open(SQL_OUT, "w", encoding="utf-8").write("\n".join(sql) + "\n")
print("\n" + "=" * 70)
print(f"TOTAL: {n_upd} updates · {n_ins} inserts · {len(sql)} sentencias -> {os.path.basename(SQL_OUT)}")
print(("\n!!! AVISOS:\n   " + "\n   ".join(warns)) if warns else "Sin avisos.")
