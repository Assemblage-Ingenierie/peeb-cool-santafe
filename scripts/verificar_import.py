# -*- coding: utf-8 -*-
"""Vérifie que la base (après import) correspond au fichier rempli, champ par champ."""
import os, sys, json, ssl, urllib.request, datetime, io
import openpyxl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # racine du repo (depuis scripts/)
FILL = os.path.join(HERE, "PEEB - Plantilla datos Admin - remplie.xlsx")

def load_env(p):
    e={}
    for l in open(p,encoding="utf-8"):
        l=l.strip()
        if l and not l.startswith("#") and "=" in l:
            k,v=l.split("=",1); e[k.strip()]=v.strip().strip('"').strip("'")
    return e
env=load_env(os.path.join(HERE,".env.local")); URL=env["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/"); KEY=env["SUPABASE_SERVICE_ROLE_KEY"]
def fetch(t,p="select=*"):
    req=urllib.request.Request(f"{URL}/rest/v1/{t}?{p}",headers={"apikey":KEY,"Authorization":f"Bearer {KEY}"})
    with urllib.request.urlopen(req,context=ssl.create_default_context()) as r: return json.loads(r.read().decode())

db={t:fetch("peebcoolsf_"+t) for t in ["documentacion_gp","gestion_financiera","eventos","equipo","entidades","capacitaciones_documentos","subproyectos","metricas","medidas","gestion_lineas"]}
ent_by_name={e["nombre"].strip().lower():e["uid"] for e in db["entidades"]}

def nt(v):
    if v is None: return None
    s=str(v).strip(); return s or None
def nb(v): return str(v).strip().lower() in ("sí","si","x","true","verdadero","1","yes","oui") if v is not None else False
def nn(v,i=False):
    if v is None or (isinstance(v,str) and not v.strip()): return None
    try: f=float(str(v).replace(",",".")) if isinstance(v,str) else float(v)
    except Exception: return None
    return int(round(f)) if i else round(f,6)
def nd(v):
    if v is None or (isinstance(v,str) and not v.strip()): return None
    if isinstance(v,(datetime.datetime,datetime.date)): return v.strftime("%Y-%m-%d")
    return str(v).strip()[:10]

T,B,I,F,D="text","bool","int","float","date"
SH=[
 ("GP - Documentación","documentacion_gp",["uid"],None,None,[("uid",T),("nombre_documento",T),("componente",T),("url",T),("confidencial",B),("publicar",B)]),
 ("Equipo - Personas","equipo",["uid"],None,None,[("uid",T),("apellido",T),("nombre",T),("entidad_uid","ent"),("rol",T),("componente",T),("telefono",T),("mail",T),("sexo",T)]),
 ("Equipo - Entidades","entidades",["uid"],None,None,[("uid",T),("nombre",T)]),
 ("Subproyectos - Medidas","medidas",["subproyecto_uid","medida"],None,None,[("subproyecto_uid",T),(None,T),("medida",T),(None,T),("activa",B),("texto",T),("kwh_anual",F)]),
 ("Subproyectos - Factibilidad","metricas",["subproyecto_uid","escenario"],"faisabilidad",None,[("subproyecto_uid",T),(None,T),("demanda_kwh",F),("demanda_despues_kwh",F),("gei_antes_tco2",F),("gei_despues_tco2",F),("costo_ee_eur",F),("costo_otras_eur",F),("benef_personal",I),("benef_personal_pct_muj",F),("benef_usuarios",I),("benef_usuarios_pct_muj",F),("benef_indirectos",I),("benef_indirectos_pct_muj",F)]),
 ("Subproyectos - Proyecto","metricas",["subproyecto_uid","escenario"],"proyecto",None,[("subproyecto_uid",T),(None,T),("demanda_kwh",F),("demanda_despues_kwh",F),("gei_antes_tco2",F),("gei_despues_tco2",F),("costo_ee_eur",F),("costo_otras_eur",F)]),
]
def conv(kind,v):
    if kind==B: return nb(v)
    if kind==I: return nn(v,True)
    if kind==F: return nn(v)
    if kind==D: return nd(v)
    if kind=="ent":
        t=nt(v); return ent_by_name.get(t.lower()) if t else None
    return nt(v)

def dbval(row,fld):
    v=row.get(fld)
    if isinstance(v,bool): return v
    if isinstance(v,(int,float)): return round(float(v),6) if not isinstance(v,bool) else v
    if isinstance(v,str): return v.strip() or None
    return v

wb=openpyxl.load_workbook(FILL,data_only=True)
total_chk=total_bad=0
for name,table,keyf,escen,_,cols in SH:
    ws=wb[name]
    idx={}
    for r in db[table]:
        if escen and r.get("escenario")!=escen: continue
        k=tuple(str(r.get(x)) for x in keyf)
        idx[k]=r
    chk=bad=0; details=[]
    for i in range(3,ws.max_row+1):
        raw=[ws.cell(i,j+1).value for j in range(len(cols))]
        if all(v is None or (isinstance(v,str) and not v.strip()) for v in raw): continue
        rec={}
        for j,(fld,kind) in enumerate(cols):
            if fld is None: continue
            rec[fld]=conv(kind,raw[j])
        if escen: rec["escenario"]=escen
        k=tuple(str(rec.get(x)) for x in keyf)
        dbrow=idx.get(k)
        if not dbrow:
            bad+=1; details.append(f"{k}: AUSENTE en BD"); continue
        for fld,kind in cols:
            if fld is None or fld in keyf or fld=="escenario": continue
            want=rec.get(fld)
            got=conv(kind,dbrow.get(fld)) if kind in (B,) else dbval(dbrow,fld)
            if kind=="ent": got=dbrow.get("entidad_uid")
            if kind==F: got=nn(dbrow.get(fld))
            if kind==I: got=nn(dbrow.get(fld),True)
            if kind==B: got=bool(dbrow.get(fld))
            if want!=got:
                bad+=1; details.append(f"{k} · {fld}: Excel={want!r} BD={got!r}")
        chk+=1
    total_chk+=chk; total_bad+=bad
    flag="OK" if bad==0 else f"!!! {bad} DIF"
    print(f"{name}: {chk} filas verificadas — {flag}")
    for d in details[:15]: print("     ", d)

print("\nConteos BD:", {t:len(db[t]) for t in ["entidades","equipo","documentacion_gp","medidas"]})
act=sum(1 for m in db["medidas"] if m.get("activa")); kw=sum(1 for m in db["medidas"] if m.get("kwh_anual") is not None)
print(f"medidas activas={act} · con kWh={kw}")
print(f"\nTOTAL: {total_chk} filas verificadas, {total_bad} diferencias.")
print("✅ La base coincide con el Excel." if total_bad==0 else "❌ Hay diferencias (ver arriba).")
